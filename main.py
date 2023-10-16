import socket
from ping3 import ping
import openpyxl


def check_ping(ip):
    response_time = ping(ip)
    if response_time is not None:
        # print(f'{ip} is pingable')
        return True
    else:
        # print(f'{ip} is not pingable')
        return False


def check_port(ip, port):
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(2)
            s.connect((ip, port))
        # print(f'{port} is open for {ip}')
        return True
    except (socket.timeout, ConnectionRefusedError):
        # print(f'{port} is not open for {ip}')
        return False


srcfile = openpyxl.load_workbook('/home/singhnavneet.su/ups_sheet_update/ups.xlsx', read_only=False, keep_vba=True)
sheetname = srcfile.active
ip_col = 1
ping_col = 1
ColNames = {}
Current = 0
for COL in sheetname.iter_cols(1, sheetname.max_column):
    ColNames[COL[0].value] = Current
    Current += 1
idx = 2
response = ""
for row_cells in sheetname.iter_rows(min_row=2, max_row=10):
    ip = row_cells[ColNames['ip']].value
    if ip is not None and ip.count('.') == 3:
        if check_ping(ip) and check_port(ip, 443) and check_port(ip, 80):
            # print(ip)
            initial = sheetname.cell(row=idx, column=ColNames['ping']+1).value
            sheetname.cell(row=idx, column=ColNames['ping']+1).value = '1'
            final = sheetname.cell(row=idx, column=ColNames['ping']+1).value
            # if str(initial) != str(final):
            #     print(
            #         f"Changing ping value from {initial} to {final} for {ip}")
        else:
            initial = sheetname.cell(row=idx, column=ColNames['ping']+1).value
            sheetname.cell(row=idx, column=ColNames['ping']+1).value = '0'
            final = sheetname.cell(row=idx, column=ColNames['ping']+1).value
            # if str(initial) != str(final):
            #     print(
            #         f"Changing ping value from {initial} to {final} for {ip}")
    response += final
    response += " "
    idx += 1

print(response)
